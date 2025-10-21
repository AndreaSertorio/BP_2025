#!/usr/bin/env python3
"""
FIX PERCEPTUAL MAP ONLY
Ripristina SOLO la sezione perceptualMap basandosi esclusivamente sul foglio PERCEPTUAL dell'Excel
"""

import json
import openpyxl
from datetime import datetime

# Paths
EXCEL_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/Competizione/eco3d-complete-competitor-report-2025-10-17.xlsx'
DB_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/financial-dashboard/src/data/database.json'

def load_perceptual_data():
    """
    Carica i dati dal foglio PERCEPTUAL dell'Excel
    Colonne:
    - A: Competitor (nome completo)
    - B: Automation Level (0-10)
    - C: Technology Innovation (0-10)
    - D: Label (nome abbreviato per visualizzazione)
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['PERCEPTUAL']
    
    positions = []
    
    # Skip header (row 1), start from row 2
    for row_idx in range(2, ws.max_row + 1):
        competitor_name = ws.cell(row_idx, 1).value
        automation = ws.cell(row_idx, 2).value
        innovation = ws.cell(row_idx, 3).value
        label = ws.cell(row_idx, 4).value
        
        if competitor_name and automation is not None and innovation is not None:
            # Converte a float
            automation_val = float(automation) if automation else 0
            innovation_val = float(innovation) if innovation else 0
            
            position = {
                "id": f"pos_{row_idx - 1:03d}",  # pos_001, pos_002, etc.
                "competitorId": f"comp_{row_idx - 1:03d}",  # Matching ID
                "name": str(label) if label else str(competitor_name),
                "x": innovation_val,  # X = Technology Innovation
                "y": automation_val,  # Y = Automation Level
                "isReference": competitor_name.lower() == 'eco3d' or (label and 'eco' in label.lower() and '3d' in label.lower())
            }
            
            positions.append(position)
            
            print(f"✓ {position['name']}: Innovation={innovation_val}, Automation={automation_val}, Reference={position['isReference']}")
    
    wb.close()
    
    return positions

def update_perceptual_map():
    """
    Aggiorna SOLO la sezione perceptualMap nel database.json
    """
    print("=" * 80)
    print("FIX PERCEPTUAL MAP - BASATO SU FOGLIO EXCEL 'PERCEPTUAL'")
    print("=" * 80)
    print()
    
    # 1. Carica dati da Excel
    print("📊 Caricamento dati dal foglio PERCEPTUAL...")
    positions = load_perceptual_data()
    print(f"✓ Caricate {len(positions)} posizioni")
    print()
    
    # 2. Carica database attuale
    print("📖 Lettura database.json...")
    with open(DB_PATH, 'r', encoding='utf-8') as f:
        db = json.load(f)
    print("✓ Database caricato")
    print()
    
    # 3. Crea nuovo perceptualMap
    perceptual_map = {
        "enabled": True,
        "axes": {
            "x": {
                "label": "Innovazione Tecnologica (0-10)",
                "min": 0,
                "max": 10,
                "minLabel": "0",
                "maxLabel": "10"
            },
            "y": {
                "label": "Livello Automazione (0-10)",
                "min": 0,
                "max": 10,
                "minLabel": "0",
                "maxLabel": "10"
            }
        },
        "positions": positions,
        "clusters": [],
        "opportunities": [
            {
                "id": "opp_high_auto_high_innov",
                "name": "High Innovation & High Automation",
                "description": "White space: alta innovazione + alta automazione (dove si posiziona Eco 3D)",
                "x": 9,
                "y": 9,
                "size": 2,
                "color": "#10b981"
            }
        ]
    }
    
    # 4. Aggiorna database
    if 'competitorAnalysis' not in db:
        print("❌ ERRORE: competitorAnalysis non trovato nel database!")
        return False
    
    if 'frameworks' not in db['competitorAnalysis']:
        db['competitorAnalysis']['frameworks'] = {}
    
    db['competitorAnalysis']['frameworks']['perceptualMap'] = perceptual_map
    
    # 5. Aggiorna metadata
    if 'metadata' in db['competitorAnalysis']:
        db['competitorAnalysis']['metadata']['lastUpdate'] = datetime.now().isoformat() + 'Z'
        db['competitorAnalysis']['metadata']['version'] = "1.1-PERCEPTUAL-FIX"
    
    # 6. Salva database
    print("💾 Salvataggio database.json...")
    with open(DB_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, indent=2, ensure_ascii=False)
    print("✓ Database aggiornato con successo!")
    print()
    
    # 7. Validazione
    print("=" * 80)
    print("VALIDAZIONE PERCEPTUAL MAP")
    print("=" * 80)
    print(f"✓ Positions: {len(positions)}")
    print(f"✓ X axis: {perceptual_map['axes']['x']['label']}")
    print(f"✓ Y axis: {perceptual_map['axes']['y']['label']}")
    print(f"✓ Enabled: {perceptual_map['enabled']}")
    
    # Trova Eco 3D
    eco3d_pos = next((p for p in positions if p['isReference']), None)
    if eco3d_pos:
        print(f"✓ Eco 3D trovato: Innovation={eco3d_pos['x']}, Automation={eco3d_pos['y']}")
    else:
        print("⚠️  Eco 3D non trovato come reference!")
    
    print()
    print("=" * 80)
    print("✅ FIX COMPLETATO!")
    print("=" * 80)
    print()
    print("🔄 PROSSIMI PASSI:")
    print("1. Ricarica il browser (Cmd+R)")
    print("2. Vai a Competitor Analysis > Perceptual Map")
    print("3. Verifica che i dati corrispondano all'Excel")
    print()
    
    return True

if __name__ == '__main__':
    success = update_perceptual_map()
    exit(0 if success else 1)
