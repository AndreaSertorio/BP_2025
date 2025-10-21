#!/usr/bin/env python3
"""
FIX PERCEPTUAL MAP - DATI ESATTI DAL FOGLIO PERCEPTUAL DELL'EXCEL
Basato ESCLUSIVAMENTE su:
- Foglio PERCEPTUAL dell'Excel
- X axis: Technology Innovation (0-10) - Colonna C
- Y axis: Automation Level (0-10) - Colonna B
- Label: Short Name - Colonna D
- Color: Basato su Tier da PROFILES
- Size: Basato su Tier
"""

import json
import openpyxl
from datetime import datetime

# Paths
EXCEL_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/Competizione/eco3d-complete-competitor-report-2025-10-17.xlsx'
DB_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/financial-dashboard/src/data/database.json'

def get_tier_color(tier):
    """Get color based on tier"""
    if tier == 'tier1':
        return '#EF4444'  # Rosso - Direct competitors
    elif tier == 'tier2':
        return '#F59E0B'  # Arancione - Indirect competitors
    else:
        return '#6B7280'  # Grigio - Patents/Other

def get_tier_size(tier):
    """Get size based on tier"""
    if tier == 'tier1':
        return 20
    elif tier == 'tier2':
        return 10
    else:
        return 5

def load_data_from_excel():
    """
    Carica dati dal foglio PERCEPTUAL dell'Excel:
    - Colonna A: Competitor (nome completo)
    - Colonna B: Automation Level (0-10)
    - Colonna C: Technology Innovation (0-10)
    - Colonna D: Label (nome abbreviato)
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Load PERCEPTUAL - UNICA FONTE DATI
    ws_perceptual = wb['PERCEPTUAL']
    
    # Load PROFILES solo per tier (colori)
    ws_profiles = wb['PROFILES']
    profiles_data = {}
    for row_idx in range(2, ws_profiles.max_row + 1):
        name = ws_profiles.cell(row_idx, 1).value
        if name:
            profiles_data[name] = {
                'tier': ws_profiles.cell(row_idx, 3).value,
                'type': ws_profiles.cell(row_idx, 4).value,
            }
    
    positions = []
    
    for row_idx in range(2, ws_perceptual.max_row + 1):
        competitor_name = ws_perceptual.cell(row_idx, 1).value  # Colonna A
        automation = ws_perceptual.cell(row_idx, 2).value      # Colonna B
        innovation = ws_perceptual.cell(row_idx, 3).value      # Colonna C
        label = ws_perceptual.cell(row_idx, 4).value           # Colonna D
        
        if competitor_name and automation is not None and innovation is not None:
            # Get tier from profiles per colore
            tier = 'tier3'  # default
            comp_type = 'substitute'  # default
            if competitor_name in profiles_data:
                tier = profiles_data[competitor_name]['tier'] or 'tier3'
                comp_type = profiles_data[competitor_name]['type'] or 'substitute'
            
            # Determina se è reference (Eco 3D)
            is_reference = 'eco3d' in str(competitor_name).lower().replace(' ', '').replace('-', '')
            
            position = {
                'id': f'pos_{row_idx - 1:03d}',
                'competitorId': f'comp_{row_idx - 1:03d}',
                'name': str(label) if label else str(competitor_name),
                'x': float(innovation) if innovation else 0,  # X = Technology Innovation
                'y': float(automation) if automation else 0,  # Y = Automation Level
                'label': str(label) if label else str(competitor_name),
                'size': 5 if is_reference else get_tier_size(tier),
                'isReference': is_reference,
                'color': '#00D2FF' if is_reference else get_tier_color(tier)
            }
            
            positions.append(position)
            
            tier_symbol = '🔵' if is_reference else ('🔴' if tier == 'tier1' else ('🟠' if tier == 'tier2' else '⚫'))
            print(f'{tier_symbol} {position["label"]}: Innovation={position["x"]:.1f}, Automation={position["y"]:.1f}')
    
    wb.close()
    return positions

def update_perceptual_map():
    """
    Aggiorna SOLO la sezione perceptualMap nel database.json
    Basato su dati ESATTI dal foglio PERCEPTUAL dell'Excel
    """
    print("=" * 80)
    print("FIX PERCEPTUAL MAP - INNOVATION vs AUTOMATION (DA FOGLIO PERCEPTUAL)")
    print("=" * 80)
    print()
    
    # 1. Carica dati da Excel
    print("📊 Caricamento dati dal foglio PERCEPTUAL...")
    positions = load_data_from_excel()
    print(f"\n✓ Caricate {len(positions)} posizioni")
    print()
    
    # 2. TUTTE le posizioni sono valide (Innovation e Automation sempre presenti)
    print(f"✓ Tutte le {len(positions)} posizioni hanno dati validi")
    print()
    
    # 3. Carica database attuale
    print("📖 Lettura database.json...")
    with open(DB_PATH, 'r', encoding='utf-8') as f:
        db = json.load(f)
    print("✓ Database caricato")
    print()
    
    # 4. Crea nuovo perceptualMap con ASSI CORRETTI dall'Excel
    perceptual_map = {
        "enabled": True,
        "description": "Posizionamento strategico su Innovation vs Automation",
        "axes": {
            "x": {
                "label": "Innovazione Tecnologica (0-10)",
                "min": 0,
                "max": 10,
                "minLabel": "0",
                "maxLabel": "10"
            },
            "y": {
                "label": "Livello Automazione (0-10)",
                "min": 0,
                "max": 10,
                "minLabel": "0",
                "maxLabel": "10"
            }
        },
        "positions": positions,
        "clusters": [],
        "opportunities": [
            {
                "id": "opp_high_innovation_high_automation",
                "name": "High Innovation & High Automation",
                "description": "White space: alta innovazione + alta automazione - posizione strategica di Eco 3D",
                "x": 9,
                "y": 9,
                "size": 2,
                "color": "#10b981"
            }
        ]
    }
    
    # 5. Aggiorna database
    if 'competitorAnalysis' not in db:
        print("❌ ERRORE: competitorAnalysis non trovato nel database!")
        return False
    
    if 'frameworks' not in db['competitorAnalysis']:
        db['competitorAnalysis']['frameworks'] = {}
    
    db['competitorAnalysis']['frameworks']['perceptualMap'] = perceptual_map
    
    # 6. Aggiorna metadata
    if 'metadata' in db['competitorAnalysis']:
        db['competitorAnalysis']['metadata']['lastUpdate'] = datetime.now().isoformat() + 'Z'
        db['competitorAnalysis']['metadata']['version'] = "1.2-PERCEPTUAL-CORRECT"
    
    # 7. Salva database
    print("💾 Salvataggio database.json...")
    with open(DB_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, indent=2, ensure_ascii=False)
    print("✓ Database aggiornato con successo!")
    print()
    
    # 8. Validazione
    print("=" * 80)
    print("VALIDAZIONE PERCEPTUAL MAP")
    print("=" * 80)
    print(f"✓ Positions totali: {len(positions)}")
    print(f"✓ X axis: {perceptual_map['axes']['x']['label']}")
    print(f"✓ Y axis: {perceptual_map['axes']['y']['label']}")
    print(f"✓ Enabled: {perceptual_map['enabled']}")
    
    # Trova Eco 3D
    eco3d_pos = next((p for p in positions if p['isReference']), None)
    if eco3d_pos:
        print(f"✓ Eco 3D: Innovation={eco3d_pos['x']:.1f}, Automation={eco3d_pos['y']:.1f}, Color={eco3d_pos['color']}")
    
    # Count by tier/color
    red_count = len([p for p in positions if p['color'] == '#EF4444'])
    orange_count = len([p for p in positions if p['color'] == '#F59E0B'])
    gray_count = len([p for p in positions if p['color'] == '#6B7280'])
    cyan_count = len([p for p in positions if p['isReference']])
    
    print(f"✓ Colori: 🔴 Tier1={red_count}, 🟠 Tier2={orange_count}, ⚫ Tier3={gray_count}, 🔵 Reference={cyan_count}")
    
    print()
    print("=" * 80)
    print("✅ FIX COMPLETATO - PERCEPTUAL MAP CON DATI CORRETTI DA EXCEL!")
    print("=" * 80)
    print()
    print("🔄 PROSSIMI PASSI:")
    print("1. Ricarica il browser (Cmd+R)")
    print("2. Vai a Competitor Analysis > Perceptual Map")
    print("3. Verifica:")
    print("   - X axis: Innovazione Tecnologica (0-10)")
    print("   - Y axis: Livello Automazione (0-10)")
    print("   - Eco 3D in alto a destra (9.5, 9.5) - massima innovation + automation")
    print("   - Pallini colorati (rossi, arancioni, grigi, cyan)")
    print("   - Nomi visibili sui pallini (label)")
    print("   - Etichetta 'High Innovation & High Automation' in alto a destra")
    print()
    
    return True

if __name__ == '__main__':
    success = update_perceptual_map()
    exit(0 if success else 1)
