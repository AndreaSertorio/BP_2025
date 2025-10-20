#!/usr/bin/env python3
"""
FIX PORTER'S 5 FORCES SCORES
Corregge i punteggi da scala 1-10 (errata) a scala 1-5 (corretta)
Basato sui dati esatti dal foglio PORTER5 dell'Excel
"""

import json
import openpyxl

# Paths
EXCEL_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/Competizione/eco3d-complete-competitor-report-2025-10-17.xlsx'
DB_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/financial-dashboard/src/data/database.json'

def load_porter5_from_excel():
    """
    Carica i dati corretti dal foglio PORTER5
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['PORTER5']
    
    porter_data = {}
    
    print("=" * 80)
    print("LOADING PORTER'S 5 FORCES FROM EXCEL")
    print("=" * 80)
    print()
    
    # Row mapping
    force_mapping = {
        2: 'competitiveRivalry',
        3: 'threatNewEntrants',
        4: 'bargainingPowerSuppliers',
        5: 'bargainingPowerBuyers',
        6: 'threatSubstitutes',
        7: 'overallAttractiveness'
    }
    
    for row_idx, key in force_mapping.items():
        force_name = ws.cell(row_idx, 1).value
        level = ws.cell(row_idx, 2).value
        score = ws.cell(row_idx, 3).value
        description = ws.cell(row_idx, 4).value
        factors = ws.cell(row_idx, 5).value
        impact = ws.cell(row_idx, 6).value
        
        if row_idx == 7:  # Overall Attractiveness
            porter_data[key] = {
                'score': float(score) if score else 3.0,
                'rating': level,
                'summary': description
            }
        else:
            # Parse factors from string (split by "•")
            factors_list = []
            if factors:
                factors_list = [f.strip() for f in str(factors).split('•') if f.strip()]
            
            porter_data[key] = {
                'score': float(score) if score else 0.0,
                'level': level,
                'description': description,
                'factors': factors_list,
                'impact': impact
            }
        
        symbol = '📊' if row_idx == 7 else '🔍'
        print(f"{symbol} {force_name}: {score}/5 ({level})")
    
    wb.close()
    print()
    print(f"✓ Loaded all 5 forces + overall from Excel")
    print()
    
    return porter_data

def fix_porter5_in_database():
    """
    Aggiorna il database con i punteggi corretti dall'Excel
    """
    print("=" * 80)
    print("FIXING PORTER'S 5 FORCES IN DATABASE")
    print("=" * 80)
    print()
    
    # 1. Load correct data from Excel
    print("📊 Loading correct data from Excel...")
    porter_excel = load_porter5_from_excel()
    
    # 2. Load database
    print("📖 Reading database.json...")
    with open(DB_PATH, 'r', encoding='utf-8') as f:
        db = json.load(f)
    
    porter_db = db.get('competitorAnalysis', {}).get('frameworks', {}).get('porter5Forces', {})
    print("✓ Found porter5Forces in database")
    print()
    
    # 3. Compare and fix
    print("🔄 Comparing scores...")
    print()
    
    for key in ['competitiveRivalry', 'threatNewEntrants', 'bargainingPowerSuppliers', 
                'bargainingPowerBuyers', 'threatSubstitutes', 'overallAttractiveness']:
        
        if key in porter_db and key in porter_excel:
            old_score = porter_db[key].get('score', 0)
            new_score = porter_excel[key]['score']
            
            if old_score != new_score:
                # Update score
                porter_db[key]['score'] = new_score
                
                # Update other fields if they exist
                for field in ['level', 'description', 'factors', 'impact', 'rating', 'summary']:
                    if field in porter_excel[key]:
                        porter_db[key][field] = porter_excel[key][field]
                
                print(f"✅ {key}: {old_score} → {new_score} (fixed!)")
            else:
                print(f"➖ {key}: {new_score} (already correct)")
    
    print()
    
    # 4. Save database
    print("💾 Saving corrected database...")
    with open(DB_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, indent=2, ensure_ascii=False)
    print("✓ Database saved successfully!")
    print()
    
    # 5. Validation
    print("=" * 80)
    print("VALIDATION - PORTER'S 5 FORCES SCORES")
    print("=" * 80)
    print()
    
    force_names = {
        'competitiveRivalry': 'Rivalry Among Competitors',
        'threatNewEntrants': 'Threat of New Entrants',
        'bargainingPowerSuppliers': 'Power of Suppliers',
        'bargainingPowerBuyers': 'Power of Buyers',
        'threatSubstitutes': 'Threat of Substitutes',
        'overallAttractiveness': 'Overall Attractiveness'
    }
    
    for key, name in force_names.items():
        if key in porter_db:
            score = porter_db[key].get('score', 0)
            level = porter_db[key].get('level') or porter_db[key].get('rating', '')
            print(f"  {name}: {score}/5 ({level})")
    
    print()
    print("=" * 80)
    print("✅ PORTER'S 5 FORCES FIXED!")
    print("=" * 80)
    print()
    print("🔄 NEXT STEPS:")
    print("1. Ricarica il browser (Cmd+R)")
    print("2. Vai a Competitor Analysis > Porter's 5")
    print("3. Verifica che tutti i score siano su scala 1-5")
    print("4. L'Overall dovrebbe essere 3.0/5 (non 6.0/5!)")
    print()
    
    return True

if __name__ == '__main__':
    success = fix_porter5_in_database()
    exit(0 if success else 1)
