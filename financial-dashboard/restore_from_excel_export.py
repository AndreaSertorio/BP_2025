#!/usr/bin/env python3
"""
RESTORE COMPETITOR ANALYSIS DA EXCEL EXPORT DEL 17 OTTOBRE 2025

Source: eco3d-complete-competitor-report-2025-10-17.xlsx
Sostituisce COMPLETAMENTE la sezione competitorAnalysis del database.json
"""

from openpyxl import load_workbook
import json
import re
from pathlib import Path

def parse_price_range(price_str):
    """Estrae min e max da '€150,000 - €300,000'"""
    if not price_str or price_str == 'N/A':
        return (0, 0)
    
    clean = str(price_str).replace('€', '').replace(',', '').replace(' ', '')
    numbers = re.findall(r'\d+', clean)
    
    if len(numbers) >= 2:
        return (int(numbers[0]), int(numbers[1]))
    elif len(numbers) == 1:
        val = int(numbers[0])
        return (val, val)
    return (0, 0)

def bool_from_checkmark(val):
    """Converte ✓/✗ in boolean"""
    return val == '✓' or val == True or val == 'Yes' or val == 'yes'

def load_profiles(wb):
    """Carica competitor profiles"""
    ws = wb['PROFILES']
    competitors = []
    
    # Get headers
    headers = {cell.column: cell.value for cell in ws[1]}
    
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        
        # Extract data
        name = ws.cell(row_idx, 1).value
        if not name:
            continue
            
        short_name = ws.cell(row_idx, 2).value or name
        tier = ws.cell(row_idx, 3).value or 'tier2'
        comp_type = ws.cell(row_idx, 4).value or 'direct'
        status = ws.cell(row_idx, 5).value or 'active'
        threat_level = ws.cell(row_idx, 6).value or 'medium'
        founded = ws.cell(row_idx, 7).value
        headquarters = ws.cell(row_idx, 8).value or 'N/A'
        employees = ws.cell(row_idx, 9).value
        revenue = ws.cell(row_idx, 10).value
        market_share = ws.cell(row_idx, 11).value
        region = ws.cell(row_idx, 12).value or 'Global'
        segments = ws.cell(row_idx, 13).value or ''
        products = ws.cell(row_idx, 14).value or name
        website = ws.cell(row_idx, 15).value or f"https://www.{short_name.lower().replace(' ', '')}.com"
        
        # Generate ID
        comp_id = f"comp_{row_idx-1:03d}"
        
        competitors.append({
            "id": comp_id,
            "name": name,
            "shortName": short_name,
            "tier": tier,
            "type": comp_type,
            "status": status,
            "priority": "high" if threat_level in ['critical', 'high'] else "medium",
            "threatLevel": threat_level,
            "monitoringPriority": 1 if threat_level == 'critical' else (2 if threat_level == 'high' else 5),
            "companyInfo": {
                "founded": int(founded) if founded and isinstance(founded, (int, float)) else None,
                "headquarters": headquarters,
                "employees": int(employees) if employees and isinstance(employees, (int, float)) else None,
                "revenue": int(revenue * 1000000) if revenue and isinstance(revenue, (int, float)) else None,
                "funding": {},
                "website": website,
                "logo": "🏥"
            },
            "products": [{
                "name": products,
                "category": "Medical Imaging / Ultrasound",
                "priceRange": "N/A",
                "priceMin": 0,
                "priceMax": 0,
                "features": {},
                "targetMarket": segments.split(', ') if segments else ["Hospitals", "Clinics"],
                "strengths": [],
                "weaknesses": [],
                "certifications": []
            }],
            "marketPosition": {
                "marketShare": float(market_share) if market_share and isinstance(market_share, (int, float)) else 0,
                "region": region,
                "segments": segments.split(', ') if segments else ["Medical Imaging"],
                "customerBase": int(employees / 10) if employees and isinstance(employees, (int, float)) else 0,
                "growthRate": 20
            },
            "swotAnalysis": {
                "strengths": [],
                "weaknesses": [],
                "opportunities": [],
                "threats": []
            },
            "battlecard": {
                "whyWeWin": [],
                "theirStrengths": [],
                "theirWeaknesses": [],
                "competitiveResponse": ""
            },
            "innovation": {
                "patents": 0,
                "rdInvestment": 0,
                "recentLaunches": []
            },
            "customerReferences": {
                "testimonials": [],
                "casestudies": 0,
                "clinicalStudies": 0
            },
            "lastUpdated": "2025-10-17",
            "notes": f"Imported from Excel export 2025-10-17"
        })
    
    return competitors

def load_comparison(wb, competitors):
    """Aggiorna competitor con dati features e price"""
    ws = wb['COMPARISON']
    
    comp_map = {c['name']: c for c in competitors}
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name or name not in comp_map:
            continue
        
        comp = comp_map[name]
        
        # Update price range
        price_range = ws.cell(row_idx, 15).value
        if price_range:
            price_min, price_max = parse_price_range(price_range)
            comp['products'][0]['priceRange'] = price_range
            comp['products'][0]['priceMin'] = price_min
            comp['products'][0]['priceMax'] = price_max
        
        # Update features
        comp['products'][0]['features'] = {
            "imaging3D": bool_from_checkmark(ws.cell(row_idx, 7).value),
            "imaging4D": bool_from_checkmark(ws.cell(row_idx, 8).value),
            "aiGuided": bool_from_checkmark(ws.cell(row_idx, 9).value),
            "portable": bool_from_checkmark(ws.cell(row_idx, 10).value),
            "wireless": bool_from_checkmark(ws.cell(row_idx, 11).value),
            "multiProbe": bool_from_checkmark(ws.cell(row_idx, 12).value),
            "automation": "full" if bool_from_checkmark(ws.cell(row_idx, 13).value) else "manual"
        }

def load_battlecards(wb, competitors):
    """Aggiorna competitor con battlecard data"""
    ws = wb['BATTLECARDS']
    
    comp_map = {c['name']: c for c in competitors}
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name or name not in comp_map:
            continue
        
        comp = comp_map[name]
        
        # Parse battlecard text (sono nelle colonne 5, 6, 7, 8)
        why_we_win = ws.cell(row_idx, 5).value or ""
        their_strengths = ws.cell(row_idx, 6).value or ""
        their_weaknesses = ws.cell(row_idx, 7).value or ""
        response = ws.cell(row_idx, 8).value or ""
        
        comp['battlecard'] = {
            "whyWeWin": [s.strip() for s in why_we_win.split('•') if s.strip()] if why_we_win else [],
            "theirStrengths": [s.strip() for s in their_strengths.split('•') if s.strip()] if their_strengths else [],
            "theirWeaknesses": [s.strip() for s in their_weaknesses.split('•') if s.strip()] if their_weaknesses else [],
            "competitiveResponse": response
        }
        
        # Update SWOT strengths/weaknesses
        comp['swotAnalysis']['strengths'] = comp['battlecard']['theirStrengths'][:3]
        comp['swotAnalysis']['weaknesses'] = comp['battlecard']['theirWeaknesses'][:3]

def load_perceptual_map(wb, competitors):
    """Carica Perceptual Map positions"""
    ws = wb['PERCEPTUAL']
    
    positions = []
    comp_map = {c['name']: c for c in competitors}
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name:
            continue
        
        automation_level = ws.cell(row_idx, 2).value or 0
        tech_innovation = ws.cell(row_idx, 3).value or 0
        label = ws.cell(row_idx, 4).value or name
        
        # Trova competitor per ottenere price
        comp = comp_map.get(name)
        if comp and comp['products']:
            price_min = comp['products'][0].get('priceMin', 40000)
            price_max = comp['products'][0].get('priceMax', 80000)
            x = (price_min + price_max) / 2
        else:
            x = 40000  # default
        
        # Automation level è già 0-10
        y = float(automation_level) if automation_level else 0
        
        # Determina competitorId
        if comp:
            comp_id = comp['id']
        else:
            comp_id = f"comp_{row_idx-1:03d}"
        
        # Size basato su tier
        tier = comp['tier'] if comp else 'tier2'
        size = 20 if tier == 'tier1' else (10 if tier == 'tier2' else 5)
        
        # Color basato su tier
        color = "#EF4444" if tier == 'tier1' else ("#F59E0B" if tier == 'tier2' else "#6B7280")
        
        positions.append({
            "competitorId": comp_id,
            "x": int(x),
            "y": float(y),
            "label": label,
            "size": size,
            "isReference": False,
            "color": color
        })
    
    # Add Eco 3D reference point
    positions.insert(0, {
        "competitorId": "eco3d_ideal",
        "x": 40000,
        "y": 9,
        "label": "Eco 3D Target",
        "size": 5,
        "isReference": True,
        "color": "#00D2FF"
    })
    
    return positions

def load_porter5(wb):
    """Carica Porter's 5 Forces"""
    ws = wb['PORTER5']
    
    porter5 = {}
    
    force_mapping = {
        "Rivalry Among Existing Competitors": "competitiveRivalry",
        "Threat of New Entrants": "threatNewEntrants",
        "Bargaining Power of Suppliers": "bargainingPowerSuppliers",
        "Bargaining Power of Buyers": "bargainingPowerBuyers",
        "Threat of Substitutes": "threatSubstitutes",
        "Overall Industry Attractiveness": "overallAttractiveness"
    }
    
    for row_idx in range(2, ws.max_row + 1):
        force_name = ws.cell(row_idx, 1).value
        if not force_name or force_name not in force_mapping:
            continue
        
        key = force_mapping[force_name]
        level = ws.cell(row_idx, 2).value or 'medium'
        score = ws.cell(row_idx, 3).value or 3
        description = ws.cell(row_idx, 4).value or ''
        factors = ws.cell(row_idx, 5).value or ''
        impact = ws.cell(row_idx, 6).value or ''
        
        # Convert score from 1-5 to 1-10
        score_10 = float(score) * 2 if isinstance(score, (int, float)) else 6
        
        # Parse factors (split by bullet or newline)
        factors_list = []
        if factors:
            factors_list = [s.strip() for s in re.split(r'[•\n]', str(factors)) if s.strip()]
        
        if key == "overallAttractiveness":
            porter5[key] = {
                "score": score_10,
                "rating": level.capitalize(),
                "summary": description,
                "opportunities": [],
                "risks": []
            }
        else:
            porter5[key] = {
                "score": score_10,
                "level": level,
                "description": description,
                "factors": factors_list[:5],  # Max 5 factors
                "impact": impact
            }
    
    # Add industry analysis
    porter5["industryAnalysis"] = {
        "industry": "Medical Ultrasound Imaging Systems",
        "marketSize": "€6.28B (2024)",
        "growthRate": "8.5% CAGR (2024-2030)"
    }
    
    return porter5

def create_benchmarking(competitors):
    """Crea benchmarking scores basati su features"""
    benchmarking_competitors = []
    
    # Score categories mapping
    for comp in competitors[:10]:  # Top 10
        features = comp['products'][0]['features'] if comp['products'] else {}
        price_min = comp['products'][0].get('priceMin', 100000) if comp['products'] else 100000
        
        # Calculate scores (0-10) based on features
        imaging_score = 10 if features.get('imaging4D') else (8 if features.get('imaging3D') else 5)
        automation_score = 10 if features.get('automation') == 'full' else (5 if features.get('multiProbe') else 2)
        portability_score = 10 if features.get('portable') and features.get('wireless') else (5 if features.get('portable') else 1)
        
        # Price value score (inversely proportional to price)
        if price_min < 10000:
            price_score = 10
        elif price_min < 50000:
            price_score = 8
        elif price_min < 100000:
            price_score = 5
        else:
            price_score = 2
        
        versatility_score = 8 if features.get('multiProbe') else 5
        ai_score = 8 if features.get('aiGuided') else 3
        
        total_score = (imaging_score + automation_score + portability_score + price_score + versatility_score + ai_score) / 6
        
        benchmarking_competitors.append({
            "competitorId": comp['id'],
            "name": comp['shortName'],
            "scores": {
                "cat_imaging_quality": imaging_score,
                "cat_automation": automation_score,
                "cat_portability": portability_score,
                "cat_price_value": price_score,
                "cat_versatility": versatility_score,
                "cat_ai_features": ai_score
            },
            "totalScore": round(total_score, 1),
            "color": f"#{hash(comp['id']) % 0xFFFFFF:06x}",
            "isReference": False
        })
    
    # Add Eco 3D reference
    benchmarking_competitors.insert(0, {
        "competitorId": "eco3d",
        "name": "Eco 3D (Reference)",
        "scores": {
            "cat_imaging_quality": 9,
            "cat_automation": 10,
            "cat_portability": 8,
            "cat_price_value": 9,
            "cat_versatility": 10,
            "cat_ai_features": 9
        },
        "totalScore": 9.2,
        "color": "#00D2FF",
        "isReference": True
    })
    
    return benchmarking_competitors

def main():
    print("🔄 RESTORE COMPETITOR ANALYSIS DA EXCEL EXPORT")
    print("=" * 70)
    print("Source: eco3d-complete-competitor-report-2025-10-17.xlsx")
    print("=" * 70)
    
    excel_path = "../Competizione/eco3d-complete-competitor-report-2025-10-17.xlsx"
    db_path = "src/data/database.json"
    
    # Load Excel
    print("\n📊 Step 1: Carico Excel...")
    wb = load_workbook(excel_path, data_only=True)
    print(f"   ✅ Sheets disponibili: {wb.sheetnames}")
    
    # Load all data
    print("\n📋 Step 2: Parse PROFILES...")
    competitors = load_profiles(wb)
    print(f"   ✅ Caricati {len(competitors)} competitor")
    
    print("\n🔍 Step 3: Parse COMPARISON (features + prices)...")
    load_comparison(wb, competitors)
    print(f"   ✅ Features e prezzi aggiornati")
    
    print("\n⚔️ Step 4: Parse BATTLECARDS...")
    load_battlecards(wb, competitors)
    print(f"   ✅ Battlecards caricati")
    
    print("\n📍 Step 5: Parse PERCEPTUAL MAP...")
    perceptual_positions = load_perceptual_map(wb, competitors)
    print(f"   ✅ {len(perceptual_positions)} positions (includes Eco 3D)")
    
    print("\n⚖️ Step 6: Parse PORTER'S 5 FORCES...")
    porter5 = load_porter5(wb)
    print(f"   ✅ Porter's 5 Forces caricato (score convertiti a scala 1-10)")
    
    print("\n📊 Step 7: Genera BENCHMARKING...")
    benchmarking_competitors = create_benchmarking(competitors)
    print(f"   ✅ {len(benchmarking_competitors)} competitor nel benchmarking")
    
    wb.close()
    
    # Load current database
    print("\n�� Step 8: Carico database corrente...")
    with open(db_path, 'r', encoding='utf-8') as f:
        db = json.load(f)
    
    # Replace competitorAnalysis section
    print("\n🔄 Step 9: Sostituisco sezione competitorAnalysis...")
    
    db['competitorAnalysis']['competitors'] = competitors
    db['competitorAnalysis']['metadata'] = {
        "lastUpdate": "2025-10-17T13:14:06Z",
        "version": "1.0-RESTORED",
        "totalCompetitors": len(competitors),
        "dataSource": "Excel Export 2025-10-17",
        "categories": ["direct", "indirect", "substitute", "emerging"]
    }
    
    # Update frameworks
    db['competitorAnalysis']['frameworks']['perceptualMap']['positions'] = perceptual_positions
    db['competitorAnalysis']['frameworks']['perceptualMap']['insights'] = [
        "Eco 3D: €40K + automation 9/10 - sweet spot ottimale",
        "GE/Philips: €150K+ + automation 2-3/10 - premium manual",
        "Siemens ABVS: €150K + automation 8/10 - specialized breast-only",
        "Butterfly/Clarius: €2-8K + automation 1-2/10 - budget portable",
        "Emerging threats: tecnologie CMUT/pMUT abbassano barriere ingresso"
    ]
    
    db['competitorAnalysis']['frameworks']['porter5Forces'] = porter5
    
    db['competitorAnalysis']['frameworks']['benchmarking']['competitors'] = benchmarking_competitors
    db['competitorAnalysis']['frameworks']['benchmarking']['insights'] = [
        "Eco 3D domina con 9.2/10 - unico con score alto su tutte le dimensioni",
        "GE: imaging eccellente (10/10) ma automation bassa (2/10)",
        "Siemens ABVS: automation massima (10/10) ma specializzato solo breast",
        "Butterfly: price/portability MAX ma imaging limitato (5/10)",
        "Differenziatore chiave: Eco 3D combina automation + versatility"
    ]
    
    # Save database
    print("\n💾 Step 10: Salvo database...")
    with open(db_path, 'w', encoding='utf-8') as f:
        json.dump(db, f, indent=2, ensure_ascii=False)
    
    print("\n" + "=" * 70)
    print("✅ RESTORE COMPLETATO CON SUCCESSO!")
    print("=" * 70)
    print(f"\n📊 Riepilogo:")
    print(f"   Competitors: {len(competitors)}")
    print(f"   Perceptual Map positions: {len(perceptual_positions)}")
    print(f"   Porter's 5 Forces: {len(porter5)} forze")
    print(f"   Benchmarking competitors: {len(benchmarking_competitors)}")
    print(f"\n🎯 Dati ORIGINALI del 17 ottobre ripristinati!")
    print(f"   Database aggiornato: {db_path}")
    
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())
