#!/usr/bin/env python3
"""
RESTORE COMPETITOR ANALYSIS - FIXED VERSION
Corregge i problemi:
1. SWOT con Opportunities e Threats (parsing corretto)
2. Perceptual Map X = PRICE (non tech innovation)
3. Porter5 nella struttura corretta per la UI
"""

from openpyxl import load_workbook
import json
import re

def parse_price_range(price_str):
    """Estrae min e max da '€150,000 - €300,000'"""
    if not price_str or price_str == 'N/A':
        return (0, 0)
    
    clean = str(price_str).replace('€', '').replace(',', '').replace(' ', '')
    numbers = re.findall(r'\d+', clean)
    
    if len(numbers) >= 2:
        return (int(numbers[0]), int(numbers[1]))
    elif len(numbers) == 1:
        val = int(numbers[0])
        return (val, val)
    return (0, 0)

def bool_from_checkmark(val):
    """Converte ✓/✗ in boolean"""
    return val == '✓' or val == True or val == 'Yes' or val == 'yes'

def load_swot_correctly(wb, competitors):
    """
    SWOT sheet ha formato:
    Competitor | Category | Items
    GE... | Strengths | Item1 | Item2 | Item3...
    GE... | Weaknesses | Item1 | Item2...
    GE... | Opportunities | Item1 | Item2...
    GE... | Threats | Item1 | Item2...
    """
    ws = wb['SWOT']
    comp_map = {c['name']: c for c in competitors}
    
    # Parse tutte le righe
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        category = ws.cell(row_idx, 2).value
        items_str = ws.cell(row_idx, 3).value
        
        if not name or not category or not items_str:
            continue
        
        if name not in comp_map:
            continue
        
        comp = comp_map[name]
        
        # Parse items (separated by |)
        items = [item.strip() for item in str(items_str).split('|') if item.strip()]
        
        # Assign to correct SWOT category
        category_lower = category.lower()
        if 'strength' in category_lower:
            comp['swotAnalysis']['strengths'] = items
        elif 'weakness' in category_lower:
            comp['swotAnalysis']['weaknesses'] = items
        elif 'opportunit' in category_lower:
            comp['swotAnalysis']['opportunities'] = items
        elif 'threat' in category_lower:
            comp['swotAnalysis']['threats'] = items

def load_profiles(wb):
    """Carica competitor profiles"""
    ws = wb['PROFILES']
    competitors = []
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name:
            continue
            
        short_name = ws.cell(row_idx, 2).value or name
        tier = ws.cell(row_idx, 3).value or 'tier2'
        comp_type = ws.cell(row_idx, 4).value or 'direct'
        status = ws.cell(row_idx, 5).value or 'active'
        threat_level = ws.cell(row_idx, 6).value or 'medium'
        founded = ws.cell(row_idx, 7).value
        headquarters = ws.cell(row_idx, 8).value or 'N/A'
        employees = ws.cell(row_idx, 9).value
        revenue = ws.cell(row_idx, 10).value
        market_share = ws.cell(row_idx, 11).value
        region = ws.cell(row_idx, 12).value or 'Global'
        segments = ws.cell(row_idx, 13).value or ''
        products = ws.cell(row_idx, 14).value or name
        website = ws.cell(row_idx, 15).value or f"https://www.{short_name.lower().replace(' ', '')}.com"
        
        comp_id = f"comp_{row_idx-1:03d}"
        
        competitors.append({
            "id": comp_id,
            "name": name,
            "shortName": short_name,
            "tier": tier,
            "type": comp_type,
            "status": status,
            "priority": "high" if threat_level in ['critical', 'high'] else "medium",
            "threatLevel": threat_level,
            "monitoringPriority": 1 if threat_level == 'critical' else (2 if threat_level == 'high' else 5),
            "companyInfo": {
                "founded": int(founded) if founded and isinstance(founded, (int, float)) else None,
                "headquarters": headquarters,
                "employees": int(employees) if employees and isinstance(employees, (int, float)) else None,
                "revenue": int(revenue * 1000000) if revenue and isinstance(revenue, (int, float)) else None,
                "funding": {},
                "website": website,
                "logo": "🏥"
            },
            "products": [{
                "name": products,
                "category": "Medical Imaging / Ultrasound",
                "priceRange": "N/A",
                "priceMin": 0,
                "priceMax": 0,
                "features": {},
                "targetMarket": segments.split(', ') if segments else ["Hospitals", "Clinics"],
                "strengths": [],
                "weaknesses": [],
                "certifications": []
            }],
            "marketPosition": {
                "marketShare": float(market_share) if market_share and isinstance(market_share, (int, float)) else 0,
                "region": region,
                "segments": segments.split(', ') if segments else ["Medical Imaging"],
                "customerBase": int(employees / 10) if employees and isinstance(employees, (int, float)) else 0,
                "growthRate": 20
            },
            "swotAnalysis": {
                "strengths": [],
                "weaknesses": [],
                "opportunities": [],
                "threats": []
            },
            "battlecard": {
                "whyWeWin": [],
                "theirStrengths": [],
                "theirWeaknesses": [],
                "competitiveResponse": ""
            },
            "innovation": {
                "patents": 0,
                "rdInvestment": 0,
                "recentLaunches": []
            },
            "customerReferences": {
                "testimonials": [],
                "casestudies": 0,
                "clinicalStudies": 0
            },
            "lastUpdated": "2025-10-17",
            "notes": f"Imported from Excel export 2025-10-17"
        })
    
    return competitors

def load_comparison(wb, competitors):
    """Aggiorna competitor con features e PRICE"""
    ws = wb['COMPARISON']
    comp_map = {c['name']: c for c in competitors}
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name or name not in comp_map:
            continue
        
        comp = comp_map[name]
        
        # IMPORTANTE: Price range per Perceptual Map X axis
        price_range = ws.cell(row_idx, 15).value
        if price_range:
            price_min, price_max = parse_price_range(price_range)
            comp['products'][0]['priceRange'] = price_range
            comp['products'][0]['priceMin'] = price_min
            comp['products'][0]['priceMax'] = price_max
        
        # Features
        comp['products'][0]['features'] = {
            "imaging3D": bool_from_checkmark(ws.cell(row_idx, 7).value),
            "imaging4D": bool_from_checkmark(ws.cell(row_idx, 8).value),
            "aiGuided": bool_from_checkmark(ws.cell(row_idx, 9).value),
            "portable": bool_from_checkmark(ws.cell(row_idx, 10).value),
            "wireless": bool_from_checkmark(ws.cell(row_idx, 11).value),
            "multiProbe": bool_from_checkmark(ws.cell(row_idx, 12).value),
            "automation": "full" if bool_from_checkmark(ws.cell(row_idx, 13).value) else "manual"
        }

def load_battlecards(wb, competitors):
    """Carica battlecards"""
    ws = wb['BATTLECARDS']
    comp_map = {c['name']: c for c in competitors}
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name or name not in comp_map:
            continue
        
        comp = comp_map[name]
        
        why_we_win = ws.cell(row_idx, 5).value or ""
        their_strengths = ws.cell(row_idx, 6).value or ""
        their_weaknesses = ws.cell(row_idx, 7).value or ""
        response = ws.cell(row_idx, 8).value or ""
        
        comp['battlecard'] = {
            "whyWeWin": [s.strip() for s in why_we_win.split('•') if s.strip()] if why_we_win else [],
            "theirStrengths": [s.strip() for s in their_strengths.split('•') if s.strip()] if their_strengths else [],
            "theirWeaknesses": [s.strip() for s in their_weaknesses.split('•') if s.strip()] if their_weaknesses else [],
            "competitiveResponse": response
        }

def load_perceptual_map_FIXED(wb, competitors):
    """
    FIXED: X axis = PRICE (da COMPARISON), Y axis = Automation Level (da PERCEPTUAL)
    """
    ws = wb['PERCEPTUAL']
    positions = []
    comp_map = {c['name']: c for c in competitors}
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name:
            continue
        
        automation_level = ws.cell(row_idx, 2).value or 0  # Y axis
        label = ws.cell(row_idx, 4).value or name
        
        # X axis = PRICE (from competitor data)
        comp = comp_map.get(name)
        if comp and comp['products']:
            price_min = comp['products'][0].get('priceMin', 40000)
            price_max = comp['products'][0].get('priceMax', 80000)
            x = (price_min + price_max) / 2
        else:
            x = 40000
        
        y = float(automation_level) if automation_level else 0
        
        comp_id = comp['id'] if comp else f"comp_{row_idx-1:03d}"
        tier = comp['tier'] if comp else 'tier2'
        size = 20 if tier == 'tier1' else (10 if tier == 'tier2' else 5)
        color = "#EF4444" if tier == 'tier1' else ("#F59E0B" if tier == 'tier2' else "#6B7280")
        
        positions.append({
            "competitorId": comp_id,
            "x": int(x),
            "y": float(y),
            "label": label,
            "size": size,
            "isReference": False,
            "color": color
        })
    
    # Eco 3D reference
    positions.insert(0, {
        "competitorId": "eco3d_ideal",
        "x": 40000,
        "y": 9,
        "label": "Eco 3D Target",
        "size": 5,
        "isReference": True,
        "color": "#00D2FF"
    })
    
    return positions

def load_porter5_FIXED(wb):
    """
    FIXED: Struttura corretta per la UI
    """
    ws = wb['PORTER5']
    
    porter5 = {}
    
    force_mapping = {
        "Rivalry Among Existing Competitors": "competitiveRivalry",
        "Threat of New Entrants": "threatNewEntrants",
        "Bargaining Power of Suppliers": "bargainingPowerSuppliers",
        "Bargaining Power of Buyers": "bargainingPowerBuyers",
        "Threat of Substitutes": "threatSubstitutes",
        "Overall Industry Attractiveness": "overallAttractiveness"
    }
    
    for row_idx in range(2, ws.max_row + 1):
        force_name = ws.cell(row_idx, 1).value
        if not force_name or force_name not in force_mapping:
            continue
        
        key = force_mapping[force_name]
        level = ws.cell(row_idx, 2).value or 'medium'
        score = ws.cell(row_idx, 3).value or 3
        description = ws.cell(row_idx, 4).value or ''
        factors = ws.cell(row_idx, 5).value or ''
        impact = ws.cell(row_idx, 6).value or ''
        
        # Convert score 1-5 → 1-10
        score_10 = float(score) * 2 if isinstance(score, (int, float)) else 6
        
        # Parse factors
        factors_list = []
        if factors:
            # Split by bullet, newline, or pipe
            factors_list = [s.strip() for s in re.split(r'[•\n|]', str(factors)) if s.strip()]
        
        if key == "overallAttractiveness":
            porter5[key] = {
                "score": score_10,
                "rating": level.capitalize(),
                "summary": description,
                "opportunities": [],
                "risks": []
            }
        else:
            porter5[key] = {
                "score": score_10,
                "level": level,
                "description": description,
                "factors": factors_list[:5],
                "impact": impact
            }
    
    # Add industry analysis
    porter5["industryAnalysis"] = {
        "industry": "Medical Ultrasound Imaging Systems",
        "marketSize": "€6.28B (2024)",
        "growthRate": "8.5% CAGR (2024-2030)"
    }
    
    # IMPORTANTE: Add enabled flag
    porter5["enabled"] = True
    
    return porter5

def create_benchmarking(competitors):
    """Crea benchmarking"""
    benchmarking_competitors = []
    
    for comp in competitors[:10]:
        features = comp['products'][0]['features'] if comp['products'] else {}
        price_min = comp['products'][0].get('priceMin', 100000) if comp['products'] else 100000
        
        imaging_score = 10 if features.get('imaging4D') else (8 if features.get('imaging3D') else 5)
        automation_score = 10 if features.get('automation') == 'full' else (5 if features.get('multiProbe') else 2)
        portability_score = 10 if features.get('portable') and features.get('wireless') else (5 if features.get('portable') else 1)
        
        if price_min < 10000:
            price_score = 10
        elif price_min < 50000:
            price_score = 8
        elif price_min < 100000:
            price_score = 5
        else:
            price_score = 2
        
        versatility_score = 8 if features.get('multiProbe') else 5
        ai_score = 8 if features.get('aiGuided') else 3
        
        total_score = (imaging_score + automation_score + portability_score + price_score + versatility_score + ai_score) / 6
        
        benchmarking_competitors.append({
            "competitorId": comp['id'],
            "name": comp['shortName'],
            "scores": {
                "cat_imaging_quality": imaging_score,
                "cat_automation": automation_score,
                "cat_portability": portability_score,
                "cat_price_value": price_score,
                "cat_versatility": versatility_score,
                "cat_ai_features": ai_score
            },
            "totalScore": round(total_score, 1),
            "color": f"#{hash(comp['id']) % 0xFFFFFF:06x}",
            "isReference": False
        })
    
    # Eco 3D reference
    benchmarking_competitors.insert(0, {
        "competitorId": "eco3d",
        "name": "Eco 3D (Reference)",
        "scores": {
            "cat_imaging_quality": 9,
            "cat_automation": 10,
            "cat_portability": 8,
            "cat_price_value": 9,
            "cat_versatility": 10,
            "cat_ai_features": 9
        },
        "totalScore": 9.2,
        "color": "#00D2FF",
        "isReference": True
    })
    
    return benchmarking_competitors

def main():
    print("🔄 RESTORE COMPETITOR ANALYSIS - FIXED VERSION")
    print("=" * 70)
    
    excel_path = "../Competizione/eco3d-complete-competitor-report-2025-10-17.xlsx"
    db_path = "src/data/database.json"
    
    wb = load_workbook(excel_path, data_only=True)
    print(f"✅ Excel caricato: {wb.sheetnames}")
    
    print("\n📋 Step 1: PROFILES...")
    competitors = load_profiles(wb)
    print(f"   ✅ {len(competitors)} competitor")
    
    print("\n💰 Step 2: COMPARISON (features + PRICES)...")
    load_comparison(wb, competitors)
    print(f"   ✅ Prices e features caricati")
    
    print("\n🎯 Step 3: SWOT (TUTTE E 4 CATEGORIE)...")
    load_swot_correctly(wb, competitors)
    # Verifica
    comp1 = competitors[0]
    print(f"   ✅ {comp1['name']}: S={len(comp1['swotAnalysis']['strengths'])}, W={len(comp1['swotAnalysis']['weaknesses'])}, O={len(comp1['swotAnalysis']['opportunities'])}, T={len(comp1['swotAnalysis']['threats'])}")
    
    print("\n⚔️ Step 4: BATTLECARDS...")
    load_battlecards(wb, competitors)
    print(f"   ✅ Battlecards caricati")
    
    print("\n📍 Step 5: PERCEPTUAL MAP (X=PRICE, Y=AUTOMATION)...")
    perceptual_positions = load_perceptual_map_FIXED(wb, competitors)
    print(f"   ✅ {len(perceptual_positions)} positions")
    print(f"      Eco 3D: x={perceptual_positions[0]['x']}, y={perceptual_positions[0]['y']}")
    print(f"      GE: x={perceptual_positions[1]['x']}, y={perceptual_positions[1]['y']}")
    
    print("\n⚖️ Step 6: PORTER'S 5 FORCES (con enabled=True)...")
    porter5 = load_porter5_FIXED(wb)
    print(f"   ✅ Porter's 5 caricato, enabled={porter5.get('enabled')}")
    
    print("\n📊 Step 7: BENCHMARKING...")
    benchmarking_competitors = create_benchmarking(competitors)
    print(f"   ✅ {len(benchmarking_competitors)} competitor")
    
    wb.close()
    
    # Load and update database
    print("\n💾 Step 8: Aggiorno database...")
    with open(db_path, 'r', encoding='utf-8') as f:
        db = json.load(f)
    
    db['competitorAnalysis']['competitors'] = competitors
    db['competitorAnalysis']['metadata'] = {
        "lastUpdate": "2025-10-17T13:14:06Z",
        "version": "1.0-FIXED",
        "totalCompetitors": len(competitors),
        "dataSource": "Excel Export 2025-10-17 (FIXED)",
        "categories": ["direct", "indirect", "substitute", "emerging"]
    }
    
    # Update frameworks
    db['competitorAnalysis']['frameworks']['perceptualMap']['enabled'] = True
    db['competitorAnalysis']['frameworks']['perceptualMap']['positions'] = perceptual_positions
    db['competitorAnalysis']['frameworks']['perceptualMap']['axes'] = {
        "x": {"label": "Prezzo (€)", "min": 0, "max": 400000, "minLabel": "€0", "maxLabel": "€400K"},
        "y": {"label": "Livello Automazione", "min": 0, "max": 10, "minLabel": "0", "maxLabel": "10"}
    }
    
    db['competitorAnalysis']['frameworks']['porter5Forces'] = porter5
    
    db['competitorAnalysis']['frameworks']['benchmarking']['enabled'] = True
    db['competitorAnalysis']['frameworks']['benchmarking']['competitors'] = benchmarking_competitors
    
    # Save
    with open(db_path, 'w', encoding='utf-8') as f:
        json.dump(db, f, indent=2, ensure_ascii=False)
    
    print("\n" + "=" * 70)
    print("✅ RESTORE COMPLETATO - FIXED VERSION!")
    print("=" * 70)
    print(f"\n📊 Validazione:")
    print(f"   Competitors: {len(competitors)}")
    print(f"   SWOT completo: Strengths={len(comp1['swotAnalysis']['strengths'])}, Weaknesses={len(comp1['swotAnalysis']['weaknesses'])}, Opportunities={len(comp1['swotAnalysis']['opportunities'])}, Threats={len(comp1['swotAnalysis']['threats'])}")
    print(f"   Perceptual Map: {len(perceptual_positions)} positions (X=PRICE)")
    print(f"   Porter's 5: enabled={porter5.get('enabled')}")
    print(f"   Benchmarking: {len(benchmarking_competitors)} competitors")
    
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())
