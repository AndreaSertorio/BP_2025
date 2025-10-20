#!/usr/bin/env python3
import openpyxl

EXCEL_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/Competizione/eco3d-complete-competitor-report-2025-10-17.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['PORTER5']

print('PORTER5 Sheet Analysis')
print('=' * 80)
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print()

print('=== HEADER ROW ===')
headers = [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
print(headers)
print()

print('=== ALL DATA ===')
for i in range(1, ws.max_row + 1):
    row_data = []
    for j in range(1, min(ws.max_column + 1, 8)):
        val = ws.cell(i, j).value
        if val:
            row_data.append(f"Col{j}: {str(val)[:100]}")
    if row_data:
        print(f'\nRow {i}:')
        for item in row_data:
            print(f'  {item}')

wb.close()
