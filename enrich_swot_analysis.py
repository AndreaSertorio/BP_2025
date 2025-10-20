#!/usr/bin/env python3
"""
ENRICH SWOT ANALYSIS - Arricchisce le sezioni SWOT dei competitor
Basato su:
- Foglio SWOT dell'Excel (source primaria)
- Competitor Eco3D.csv (differenziazioni tecniche)
- Guida framework SWOT
"""

import json
import openpyxl
import csv
from collections import defaultdict

# Paths
EXCEL_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/Competizione/eco3d-complete-competitor-report-2025-10-17.xlsx'
CSV_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/assets/Competitor Eco3D.csv'
DB_PATH = '/Users/dracs/Documents/START_UP/DOC PROGETTI/DOC_ECO 3d Multisonda/__BP 2025/financial-dashboard/src/data/database.json'

def load_swot_from_excel():
    """
    Carica SWOT da Excel - formato:
    Colonna A: Competitor
    Colonna B: Category (Strengths/Weaknesses/Opportunities/Threats)
    Colonna C: Items (separati da " | ")
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SWOT']
    
    swot_data = defaultdict(lambda: {
        'strengths': [],
        'weaknesses': [],
        'opportunities': [],
        'threats': []
    })
    
    print("=" * 80)
    print("LOADING SWOT FROM EXCEL")
    print("=" * 80)
    print()
    
    for row_idx in range(2, ws.max_row + 1):
        competitor_name = ws.cell(row_idx, 1).value
        category = ws.cell(row_idx, 2).value
        items_raw = ws.cell(row_idx, 3).value
        
        if not competitor_name or not category or not items_raw:
            continue
        
        # Split items by " | " delimiter
        items = [item.strip() for item in str(items_raw).split('|')]
        
        # Map category to SWOT key
        category_map = {
            'Strengths': 'strengths',
            'Weaknesses': 'weaknesses',
            'Opportunities': 'opportunities',
            'Threats': 'threats'
        }
        
        swot_key = category_map.get(category)
        if swot_key:
            swot_data[competitor_name][swot_key].extend(items)
    
    wb.close()
    
    print(f"✓ Loaded SWOT for {len(swot_data)} competitors from Excel")
    print()
    
    return swot_data

def load_differentiators_from_csv():
    """
    Carica differenziatori tecnici da CSV per arricchire SWOT
    """
    diff_data = {}
    
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            tech = row.get('Documento / tecnologia', '').strip()
            key_points = row.get('Punti chiave noti', '').strip()
            limits = row.get('Limite tecnico/clinico principale', '').strip()
            eco3d_diff = row.get('In cosa Eco 3D si differenzia / migliora concretamente', '').strip()
            
            if tech:
                # Parse limits into weaknesses (split by "- " or newlines)
                weaknesses = []
                if limits:
                    # Clean HTML tags and split
                    limits_clean = limits.replace('<br/>', '\n').replace('- ', '\n')
                    weaknesses = [w.strip() for w in limits_clean.split('\n') if w.strip() and len(w.strip()) > 10]
                
                # Parse key points into strengths
                strengths = []
                if key_points:
                    key_clean = key_points.replace('<br/>', '\n').replace('- ', '\n')
                    strengths = [s.strip() for s in key_clean.split('\n') if s.strip() and len(s.strip()) > 10]
                
                # Eco 3D diff becomes a threat for them
                threats = []
                if eco3d_diff:
                    eco_clean = eco3d_diff.replace('<br/>', '\n').replace('– ', '\n')
                    threats = [f"Eco 3D: {t.strip()}" for t in eco_clean.split('\n') if t.strip() and len(t.strip()) > 15]
                
                diff_data[tech] = {
                    'strengths': strengths[:3],  # Top 3
                    'weaknesses': weaknesses[:3],  # Top 3
                    'threats': threats[:2]  # Top 2
                }
    
    print(f"✓ Loaded {len(diff_data)} technical differentiators from CSV")
    print()
    
    return diff_data

def normalize_competitor_name(name):
    """
    Normalizza nome competitor per matching
    """
    # Remove common suffixes and normalize
    name = str(name).lower()
    name = name.replace(' - ', ' ').replace('-', ' ')
    name = name.replace('/', ' ')
    name = ' '.join(name.split())  # normalize spaces
    return name

def find_competitor_in_db(swot_name, db_competitors):
    """
    Trova competitor nel database basandosi sul nome SWOT
    Usa fuzzy matching
    """
    swot_norm = normalize_competitor_name(swot_name)
    
    # Direct keyword matching
    keywords = swot_norm.split()
    
    for comp in db_competitors:
        comp_name = comp.get('name', '')
        comp_norm = normalize_competitor_name(comp_name)
        
        # Check if any significant keyword matches
        matches = sum(1 for kw in keywords if kw in comp_norm and len(kw) > 3)
        if matches >= 1:  # At least one significant match
            return comp
    
    return None

def enrich_database():
    """
    Arricchisce il database con SWOT completo dall'Excel
    """
    print("=" * 80)
    print("ENRICHING DATABASE WITH SWOT ANALYSIS")
    print("=" * 80)
    print()
    
    # 1. Load data sources
    print("📊 Loading data sources...")
    swot_excel = load_swot_from_excel()
    diff_csv = load_differentiators_from_csv()
    
    # 2. Load database
    print("📖 Reading database.json...")
    with open(DB_PATH, 'r', encoding='utf-8') as f:
        db = json.load(f)
    
    competitors = db.get('competitorAnalysis', {}).get('competitors', [])
    print(f"✓ Found {len(competitors)} competitors in database")
    print()
    
    # 3. Match and enrich from Excel
    print("🔄 Matching SWOT data with database competitors...")
    print()
    
    matched = 0
    enriched = 0
    
    for swot_name, swot_data in swot_excel.items():
        comp = find_competitor_in_db(swot_name, competitors)
        
        if comp:
            matched += 1
            old_swot = comp.get('swotAnalysis', {})
            
            # Count items before
            old_total = sum(len(old_swot.get(k, [])) for k in ['strengths', 'weaknesses', 'opportunities', 'threats'])
            
            # Update SWOT with Excel data
            comp['swotAnalysis'] = swot_data
            
            # Count items after
            new_total = sum(len(swot_data.get(k, [])) for k in ['strengths', 'weaknesses', 'opportunities', 'threats'])
            
            if new_total > old_total:
                enriched += 1
                print(f"✅ {comp['name']}: {old_total} → {new_total} items (+{new_total - old_total})")
            else:
                print(f"➖ {comp['name']}: Already complete ({new_total} items)")
        else:
            print(f"⚠️  Could not match '{swot_name}' to database competitor")
    
    # 4. Enhance partial SWOT with CSV technical data
    print()
    print("🔬 Enhancing partial SWOT with CSV technical details...")
    print()
    
    csv_enhanced = 0
    for comp in competitors:
        swot = comp.get('swotAnalysis', {})
        total = sum(len(swot.get(k, [])) for k in ['strengths', 'weaknesses', 'opportunities', 'threats'])
        
        # If SWOT is partial (< 8 items), try to enhance with CSV
        if total < 8:
            comp_name = comp.get('name', '')
            
            # Try to find matching CSV entry
            for tech_name, csv_data in diff_csv.items():
                if normalize_competitor_name(comp_name) in normalize_competitor_name(tech_name) or \
                   normalize_competitor_name(tech_name) in normalize_competitor_name(comp_name):
                    
                    # Merge CSV data with existing SWOT
                    before = total
                    
                    if csv_data.get('strengths'):
                        swot.setdefault('strengths', []).extend(csv_data['strengths'])
                        swot['strengths'] = list(dict.fromkeys(swot['strengths']))  # Remove duplicates
                    
                    if csv_data.get('weaknesses'):
                        swot.setdefault('weaknesses', []).extend(csv_data['weaknesses'])
                        swot['weaknesses'] = list(dict.fromkeys(swot['weaknesses']))
                    
                    if csv_data.get('threats'):
                        swot.setdefault('threats', []).extend(csv_data['threats'])
                        swot['threats'] = list(dict.fromkeys(swot['threats']))
                    
                    comp['swotAnalysis'] = swot
                    
                    after = sum(len(swot.get(k, [])) for k in ['strengths', 'weaknesses', 'opportunities', 'threats'])
                    
                    if after > before:
                        csv_enhanced += 1
                        print(f"🔬 {comp_name}: Enhanced with CSV data ({before} → {after} items)")
                    
                    break
    
    print(f"   CSV enhanced: {csv_enhanced} competitors")
    print()
    
    print()
    print(f"📊 Summary:")
    print(f"   - Matched: {matched}/{len(swot_excel)}")
    print(f"   - Enriched: {enriched}")
    print()
    
    # 5. Save database
    print("💾 Saving enriched database...")
    with open(DB_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, indent=2, ensure_ascii=False)
    print("✓ Database saved successfully!")
    print()
    
    # 6. Validation summary
    print("=" * 80)
    print("VALIDATION - SWOT COMPLETENESS")
    print("=" * 80)
    print()
    
    for comp in competitors[:10]:  # Show first 10
        swot = comp.get('swotAnalysis', {})
        total = sum(len(swot.get(k, [])) for k in ['strengths', 'weaknesses', 'opportunities', 'threats'])
        completeness = "✅ Complete" if total >= 8 else "⚠️  Partial" if total >= 4 else "❌ Empty"
        
        print(f"{completeness} {comp['name']}: S:{len(swot.get('strengths', []))} W:{len(swot.get('weaknesses', []))} O:{len(swot.get('opportunities', []))} T:{len(swot.get('threats', []))}")
    
    print()
    print("=" * 80)
    print("✅ SWOT ENRICHMENT COMPLETED!")
    print("=" * 80)
    print()
    print("🔄 NEXT STEPS:")
    print("1. Ricarica il browser (Cmd+R)")
    print("2. Vai a Competitor Analysis > SWOT Analysis")
    print("3. Verifica che i dati siano completi")
    print()
    
    return True

if __name__ == '__main__':
    success = enrich_database()
    exit(0 if success else 1)
