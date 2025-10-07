#!/usr/bin/env python3
"""
Analizza il file Excel Costi_sviluppo.xlsx ed estrae la struttura
per creare il modulo Budget nel dashboard
"""

import openpyxl
import json
from pathlib import Path

def analyze_excel(file_path):
    """Analizza l'Excel e estrae struttura e dati"""
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"📊 File Excel caricato: {file_path}")
    print(f"📋 Fogli disponibili: {wb.sheetnames}\n")
    
    results = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"📄 Foglio: {sheet_name}")
        print(f"{'='*60}")
        
        # Dimensioni
        max_row = sheet.max_row
        max_col = sheet.max_column
        print(f"Dimensioni: {max_row} righe × {max_col} colonne")
        
        # Leggi tutte le righe
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if any(cell is not None for cell in row):  # Salta righe vuote
                row_data = {
                    'row': row_idx,
                    'values': [cell if cell is not None else '' for cell in row]
                }
                data.append(row_data)
                
                # Mostra prime 20 righe per capire la struttura
                if row_idx <= 20:
                    formatted_row = ' | '.join([str(cell)[:30] for cell in row_data['values'][:10]])
                    print(f"  R{row_idx:3d}: {formatted_row}")
        
        results[sheet_name] = {
            'dimensions': {'rows': max_row, 'cols': max_col},
            'data': data
        }
        
        print(f"\n✅ Estratte {len(data)} righe con dati")
    
    # Salva JSON per analisi
    output_file = Path(file_path).parent / 'budget_structure.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\n\n💾 Struttura salvata in: {output_file}")
    
    return results

if __name__ == '__main__':
    excel_file = Path(__file__).parent / 'Costi_sviluppo.xlsx'
    
    if not excel_file.exists():
        print(f"❌ File non trovato: {excel_file}")
        exit(1)
    
    try:
        analyze_excel(excel_file)
        print("\n✅ Analisi completata!")
    except Exception as e:
        print(f"\n❌ Errore durante l'analisi: {e}")
        import traceback
        traceback.print_exc()
